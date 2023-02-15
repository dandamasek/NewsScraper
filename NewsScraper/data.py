from openpyxl import Workbook
from web import Web


class Data(Web):
    #data_from_news = [[]]*3
    news_vector = []
    #keyword_data_from_news = [[]]*3
    keywords = []

    def __init__(self, txt_save_data):
        self.txt_save_data = txt_save_data 

    def addNewsToData(self,news):
        self.news_vector.append(news)

    def cleanData(self):
        
        f = open(str(self.txt_save_data),"r")
        lines = f.readlines()
        f.close()
        
        #delete duplicity, duplicate to none
        lines = dict.fromkeys(lines)
        #delete none from lines
        lines = list(filter(None,lines))

        #delete all lines without full information
        f = open(self.txt_save_data,"w")
        for line in lines:
            va_line = line.split(' |-| ')
            #lenght of data and second is real url
            if len(va_line) >= 4 and va_line[1].find("https:") != -1:
                f.write(line)

        f.close()

    def updateData(self):
        type_list = []
        headlines_list = []
        urls_list = []
        text_list = []
        text_help_prom = ""
        with open(str(self.txt_save_data)) as file:
            for line in file:
                line = line.split(' |-| ')
                urls_list.append(line[1])
                headlines_list.append(line[2])
                type_list.append(line[0])
                
                text_help_prom = ""
                for indexLine in range(3,len(line)-1):
                    text_help_prom += line[indexLine]

                text_list.append(text_help_prom)
                
        self.data_from_news = zip(type_list,headlines_list,urls_list,text_list)

    def writeData(self):
        for x in self.data_from_news:
            print(x)
            print('\n')          

    def saveToExcel(self):
        #check if excel adn sheets in it exists if not create it
        try:
            wb = Workbook()
            #not rewrite just append new data
            #wb = load_workbook(excelName+".xlsx")
        except:
            #wb = Workbook()
            pass
        try:
            wsS = wb['SortedNews']
        except:
            wsS = wb.active
            wsS.title = "SortedNews"

        try:
            wsA = wb['AllNews']
        except:
            wsA = wb.create_sheet('AllNews')

        f = open(self.txt_save_data,"r")
        lines = f.readlines()
        f.close()
        #all data
        for line in lines:
            split_line = line.split('|-|')
            wsA.append([split_line[0],split_line[2],split_line[1]])

        #sorted data
        for key in self.keyword_data_from_news:
            wsS.append([key[0],key[2],key[1]])

        wb.save(self.excel_save_data)


    def news_scraper_by_title(self):
        keyword_list_headline = []
        keyword_list_url = []
        keyword_list_type = []
        keyword_list_text = []
        # Goes through the list and searches fot the keyword
        for title in self.data_from_news:

            #keywords lower
            for key in self.keywords:
                key.lower()

            #find keyword
            for indKey in range (len(self.keywords)):
                if self.keywords[indKey] in title[1].lower():
                    keyword_list_type.append(title[0])
                    keyword_list_headline.append(title[1])
                    keyword_list_url.append(title[2])
                    keyword_list_text.append(title[3])

        self.keyword_data_from_news = zip(keyword_list_type,keyword_list_headline,keyword_list_url,keyword_list_text)   
        #delete duplicity, duplicate to none
        self.keyword_data_from_news = dict.fromkeys(self.keyword_data_from_news)
        #delete none from list
        self.keyword_data_from_news = list(filter(None,self.keyword_data_from_news))
     

    def news_scraper_by_text(self):
        keyword_list_headline = []
        keyword_list_url = []
        keyword_list_type = []
        keyword_list_text = []
        
        self.keyword_data_from_news = []
        # Goes through the list and searches fot the keyword
        for i, title in enumerate(self.data_from_news):
            
            keywords = ""
            #keywords lower
            for key in self.keywords:
                key.lower()

            #find keywords from txt

            for choose_keyword in self.keywords:
                if title[3].find(choose_keyword) != -1:
                    keyword_list_type.append(title[0])
                    keyword_list_headline.append(title[1])
                    keyword_list_url.append(title[2])
                    keyword_list_text.append(title[3])   

        self.keyword_data_from_news = zip(keyword_list_type,keyword_list_headline,keyword_list_url,keyword_list_text)   
        #delete duplicity, duplicate to none
        self.keyword_data_from_news = dict.fromkeys(self.keyword_data_from_news)
        #delete none from list
        self.keyword_data_from_news = list(filter(None,self.keyword_data_from_news))


        











